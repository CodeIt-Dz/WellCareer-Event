from .models import Student, FileTracker
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from rest_framework import viewsets, status
from rest_framework.response import Response
from django.conf import settings
from .serializers import StudentSerializer

class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

    def create(self, request, *args, **kwargs):
        # Save the student to the database
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        # Get the created student data
        student_data = serializer.data

        # Define the file path for the Excel file
        media_path = os.path.join(settings.MEDIA_ROOT, 'excel_files/')
        os.makedirs(media_path, exist_ok=True)  # Ensure the directory exists
        file_path = os.path.join(media_path, 'students.xlsx')

        # Check if the file exists
        if not os.path.exists(file_path):
            # Create a new workbook and add headers if the file doesn't exist
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Students"
            headers = [
                "First Name",
                "Last Name",
                "Gender",
                "Phone Number",
                "Birth Date",
                "Major",
                "Education Level",
                "Professional Interests",
                "CV File"
            ]
            sheet.append(headers)
        else:
            # Load the existing workbook
            workbook = load_workbook(file_path)
            sheet = workbook.active

        # Append the new student data
        sheet.append([
            student_data.get('first_name'),
            student_data.get('last_name'),
            student_data.get('gender'),
            student_data.get('phone_number'),
            student_data.get('birth_date'),
            student_data.get('major'),
            student_data.get('education_level'),
            ", ".join(student_data.get('professional_interests', [])),  # Convert JSON to a string
            student_data.get('cv'),  # Add CV file path
        ])

        # Save the workbook to the file
        workbook.save(file_path)

        # Save or update the file in the `FileTracker` model
        file_tracker, created = FileTracker.objects.get_or_create(id=1)  # Use a single instance with ID=1
        file_tracker.file.name = os.path.relpath(file_path, settings.MEDIA_ROOT)  # Save relative path
        file_tracker.save()

        return Response(serializer.data, status=status.HTTP_201_CREATED)
